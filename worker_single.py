import datetime
import os
import random
import shutil
import time
from pathlib import Path
from queue import Queue
from threading import Thread, Lock

import openpyxl


##№Класс, создающий файл Excel
def worker(item: {}):
    print(f"Начало работы над элементом с id {item['id']}"
          f" в {datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
    item["state"] = "in_progress"
    print(str(item) + "\r\n")

    try:
        path_to_excel = Path(os.getcwd(), "files_to_send", f"{item['id']}.xlsx")
        path_to_template = Path(os.getcwd(), "Товарный-чек.xlsx")
        shutil.copyfile(path_to_template, path_to_excel)
        wb = openpyxl.load_workbook(path_to_excel)
        ws = wb.active
        ws["A1"].value = item["org_name"]
        title_cell = ws["A4"].value
        title_cell = title_cell.replace("id", str(item["id"])).replace("date", item["date"])
        ws["A4"].value = title_cell
        all_sum = 0
        i = 6
        for product in item["product_list"]:
            ws[f"A{i}"].value = i - 5
            ws[f"B{i}"].value = product[0]
            ws[f"H{i}"].value = product[1]
            ws[f"J{i}"].value = product[2]
            ws[f"M{i}"].value = product[1] * product[2]
            all_sum += product[1] * product[2]
            i += 1
        ws["M16"].value = all_sum
        wb.save(path_to_excel)
        wb.close()
        time.sleep(5)
        item["state"] = "Done Successfully"
        print(str(item) + "\r\n")
    except Exception as e:
        print(e)

    print(f"Конец работы над элементом с id {item['id']}"
          f" в {datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")